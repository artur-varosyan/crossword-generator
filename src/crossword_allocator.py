from openpyxl.comments import Comment
from classes import Word
from definition_finder import find_definition


MAX_COLUMN = 25
MAX_ROW = 25
NO_LETTERS = "@"
NO_VERTICAL = "!"
NO_HORIZONTAL = "="


def allocate_word(temp_word, word_list, crossword, hints):
    if len(word_list) == 0:
        # If first word allocate in the centre of the crossword
        left_indent = round((25 - len(temp_word)) / 2)
        word_list.append(Word(temp_word, "horizontal", 12, left_indent, len(temp_word)))
    else:
        # Check if can match with existing words
        matched, word_list = match_words(temp_word, word_list, crossword)
        if not matched:  # New word is needed
            return

    map_word(word_list, crossword)
    definition = find_definition(word_list[-1].name)
    hints.cell(3 + len(word_list), 2).value = word_list[-1].orientation
    hints.cell(3 + len(word_list), 3).value = definition


def match_words(new_word, word_list, crossword):
    i = 0
    while i < len(word_list):
        comparison_word = word_list[len(word_list) - 1 - i].name
        for j in range(0, len(comparison_word)):
            for k in range(0, len(new_word)):
                if comparison_word[j] == new_word[k]:
                    # Matched letter, check surroundings
                    overlap, word_list = check_surroundings(new_word, word_list, j, k, crossword, i)
                    if not overlap:
                        return True, word_list
        # No match for this comparison word - try another one
        i = i + 1

    # No match AT ALL in crossword
    return False, word_list


def check_surroundings(temp_word, word_list, index, index_temp, crossword, i):
    # Identify orientation and actual position of the new word
    if word_list[len(word_list) - 1 - i].orientation == "horizontal":
        actual_row = word_list[len(word_list) - 1 - i].starting_row
        actual_column = word_list[len(word_list) - 1 - i].starting_column + index
        new_orientation = "vertical"
    else:
        actual_row = word_list[len(word_list) - 1 - i].starting_row + index
        actual_column = word_list[len(word_list) - 1 - i].starting_column
        new_orientation = "horizontal"

    # Count how many letters are before intersection and how many after
    before_intersection = index_temp
    after_intersection = len(temp_word) - index_temp - 1

    # Check for overlap with existing words in the crossword
    # Keys:
    #   = No horizontal letters allowed
    #   ! No vertical letters allowed
    #   @ no letters at all allowed
    overlap = False

    try:
        if new_orientation == "vertical":
            # ONE BEFORE First character
            if crossword.cell(actual_row - before_intersection - 1, actual_column).value is not None and crossword.cell(actual_row - before_intersection - 1, actual_column).value != NO_HORIZONTAL and crossword.cell(actual_row - before_intersection - 1, actual_column).value != NO_VERTICAL and crossword.cell(actual_row - before_intersection - 1, actual_column).value != NO_LETTERS:
                overlap = True
            # ONE AFTER First character
            if crossword.cell(actual_row + after_intersection + 1, actual_column).value is not None and crossword.cell(actual_row + after_intersection + 1, actual_column).value != NO_HORIZONTAL and crossword.cell(actual_row + after_intersection + 1, actual_column).value != NO_VERTICAL and crossword.cell(actual_row + after_intersection + 1, actual_column).value != NO_LETTERS:
                overlap = True
        else:
            # ONE BEFORE First character
            if crossword.cell(actual_row, actual_column - before_intersection - 1).value is not None and crossword.cell(actual_row, actual_column - before_intersection - 1).value != NO_HORIZONTAL and crossword.cell(actual_row, actual_column - before_intersection - 1).value != NO_VERTICAL and crossword.cell(actual_row, actual_column - before_intersection - 1).value != NO_LETTERS:
                overlap = True
            # ONE AFTER First character
            if crossword.cell(actual_row, actual_column + after_intersection + 1).value is not None and crossword.cell(actual_row, actual_column + after_intersection + 1).value != NO_HORIZONTAL and crossword.cell(actual_row, actual_column + after_intersection + 1).value != NO_VERTICAL and crossword.cell(actual_row, actual_column + after_intersection + 1).value != NO_LETTERS:
                overlap = True

        for cell in range(0, before_intersection):
            # Bug fix (shouldn't change letters now?) below
            if overlap:
                continue
            if new_orientation == "vertical":
                if (actual_row - cell) > MAX_ROW or actual_column > MAX_COLUMN:
                    overlap = True
                elif crossword.cell(actual_row - 1 - cell, actual_column).value is not None and crossword.cell(actual_row - 1 - cell, actual_column).value != NO_HORIZONTAL:
                    overlap = True
            else:
                if (actual_column - cell) > MAX_COLUMN or actual_row > MAX_ROW:
                        overlap = True
                elif crossword.cell(actual_row, actual_column - 1 - cell).value is not None and crossword.cell(actual_row, actual_column - 1 - cell).value != NO_VERTICAL:
                    overlap = True
        for cell in range(1, after_intersection + 1):
            if overlap == True:
                continue
            if new_orientation == "vertical":
                if (actual_row + cell) > MAX_ROW or actual_column > MAX_COLUMN:
                    overlap = True
                elif crossword.cell(actual_row + cell, actual_column).value is not None and crossword.cell(actual_row + cell, actual_column).value != NO_HORIZONTAL:
                    overlap = True
            else:
                if (actual_column + cell) > MAX_COLUMN or actual_row > MAX_ROW:
                    overlap = True
                if crossword.cell(actual_row, actual_column + cell).value is not None and crossword.cell(actual_row, actual_column + cell).value != NO_VERTICAL:
                    overlap = True
    except ValueError:
        # Word goes beyond the range of the crossword!
        overlap = True

    # Append the word to the list if no overlaps have been found
    if not overlap:
        if new_orientation == "vertical":
            word_list.append(
                Word(temp_word, new_orientation,
                     actual_row - before_intersection,
                     actual_column, len(temp_word)))
        else:
            word_list.append(
                Word(temp_word, new_orientation,
                     actual_row,
                     actual_column - before_intersection, len(temp_word)))

    return overlap, word_list


def map_word(word_list, crossword):
    new_word = word_list[len(word_list) - 1]
    if new_word.orientation == "vertical":
        i = 0
        temp_row = new_word.starting_row
        for letter in new_word.name:
            if i == 0:
                # Inserting comments with word id and orientation
                i = 1
                crossword.cell(temp_row, new_word.starting_column).value = letter
                comment = Comment(f"{len(word_list)} {new_word.orientation}", "CrosswordGenerator")
                if crossword.cell(temp_row, new_word.starting_column).comment is None:
                    crossword.cell(temp_row, new_word.starting_column).comment = comment
                else:
                    crossword.cell(temp_row, new_word.starting_column).comment.text = crossword.cell(temp_row, new_word.starting_column).comment.text + " " + comment.text
            else:
                crossword.cell(temp_row, new_word.starting_column).value = letter

            if temp_row == new_word.starting_row:
                map_limits(temp_row - 1, new_word.starting_column, "both", crossword)
            elif temp_row == (new_word.starting_row + new_word.length - 1):
                map_limits(temp_row + 1, new_word.starting_column, "both", crossword)

            map_limits(temp_row, new_word.starting_column - 1, "vertical", crossword)
            map_limits(temp_row, new_word.starting_column + 1, "vertical", crossword)

            temp_row = temp_row + 1
    else:
        i = 0
        temp_column = new_word.starting_column
        for letter in new_word.name:
            if i == 0:
                # Inserting comments with word id and orientation
                i = 1
                crossword.cell(new_word.starting_row, temp_column).value = letter
                comment = Comment(f"{len(word_list)} {new_word.orientation}", "CrosswordGenerator")
                if crossword.cell(new_word.starting_row, temp_column).comment is None:
                    crossword.cell(new_word.starting_row, temp_column).comment = comment
                else:
                    crossword.cell(new_word.starting_row, temp_column).comment.text = crossword.cell(new_word.starting_row, temp_column).comment.text + " " + comment.text
            else:
                crossword.cell(new_word.starting_row, temp_column).value = letter

            if temp_column == new_word.starting_column:
                map_limits(new_word.starting_row, temp_column - 1, "both", crossword)
            elif temp_column == (new_word.starting_column + new_word.length - 1):
                map_limits(new_word.starting_row, temp_column + 1, "both", crossword)

            map_limits(new_word.starting_row - 1, temp_column, "horizontal", crossword)
            map_limits(new_word.starting_row + 1, temp_column, "horizontal", crossword)

            temp_column = temp_column + 1


def map_limits(row, column, orientation, crossword):
    if orientation == "vertical":
        if crossword.cell(row, column).value is None:
            crossword.cell(row, column).value = NO_VERTICAL
        elif crossword.cell(row, column).value == NO_HORIZONTAL:
            crossword.cell(row, column).value = NO_LETTERS
    elif orientation == "horizontal":
        if crossword.cell(row, column).value is None:
            crossword.cell(row, column).value = NO_HORIZONTAL
        elif crossword.cell(row, column).value == NO_VERTICAL:
            crossword.cell(row, column).value = NO_LETTERS
    else:
        if crossword.cell(row, column).value is None:
            crossword.cell(row, column).value = NO_LETTERS
        elif crossword.cell(row, column).value == NO_VERTICAL or crossword.cell(row, column).value == NO_HORIZONTAL:
            crossword.cell(row, column).value = NO_LETTERS

