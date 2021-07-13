from openpyxl.styles import PatternFill


def format_crossword(crossword, answers):
    black_fill = PatternFill(start_color='00000000',
                             end_color='00000000',
                             fill_type='solid')
    for x in range(1, 27):
        for y in range(1, 27):
            print(crossword.cell(x, y).value)
            if crossword.cell(x, y).value is None or crossword.cell(x, y).value == "" or \
                    crossword.cell(x,y).value == "@" or crossword.cell(x, y).value == "!" or \
                    crossword.cell(x, y).value == "=":
                crossword.cell(x, y).value = ""
                crossword.cell(x, y).fill = black_fill
                answers.cell(x, y).value = ""
                answers.cell(x, y).fill = black_fill
            else:
                crossword.cell(x, y).value = ""
